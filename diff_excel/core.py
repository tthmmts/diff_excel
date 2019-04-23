from datetime import datetime as dt
from pathlib import Path

import openpyxl
import pytz
from openpyxl.styles import Font, NamedStyle, PatternFill

from . import helpers


def diff_excel_data(from_file_full_path: Path,
                    to_file_full_path: Path,
                    key: str = "") -> dict:
    if key is None:
        key = ""
    from_dict, from_keys, from_columns = helpers.get_dict_data_from_excel_file(from_file_full_path, key=key)
    to_dict, to_keys, to_columns = helpers.get_dict_data_from_excel_file(to_file_full_path, key=key)

    key_status = helpers.list_ndiff(from_keys, to_keys)
    column_status = helpers.list_ndiff(from_columns, to_columns)

    # 新旧 キー（行）名のリスト
    diff_keys = [key[2:] for key in key_status]
    # 新旧 カラム（列）名のリスト
    diff_columns = [column[2:] for column in column_status]

    diff_results = {}
    for key in key_status:
        key_c, key_v = key[:2], key[2:]
        for column in column_status:
            column_c, column_v = column[:2], column[2:]
            diff_results.setdefault(key_v, {})
            if key_c == "  " and column_c == "  ":
                if from_dict[key_v][column_v] != to_dict[key_v][column_v]:
                    value = {"code": "=>",
                             "status": "replace",
                             "from": from_dict[key_v][column_v],
                             "to": to_dict[key_v][column_v],
                             "value": to_dict[key_v][column_v]
                             }
                    diff_results[key_v].setdefault(column_v, value)
                else:
                    value = {"code": "==",
                             "status": "equal",
                             "from": from_dict[key_v][column_v],
                             "to": to_dict[key_v][column_v],
                             "value": to_dict[key_v][column_v]
                             }
                    diff_results[key_v].setdefault(column_v, value)
            elif key_c == "- " and column_c == "  ":
                value = {"code": "l-",
                         "status": "deleteLine",
                         "from": from_dict[key_v][column_v],
                         "to": "",
                         "value": from_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)
            elif key_c == "  " and column_c == "- ":
                value = {"code": "r-",
                         "status": "deleteRow",
                         "from": from_dict[key_v][column_v],
                         "to": "",
                         "value": from_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)
            elif key_c == "- " and column_c == "- ":
                value = {"code": "--",
                         "status": "delete",
                         "from": from_dict[key_v][column_v],
                         "to": "",
                         "value": from_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)
            elif key_c == "+ " and column_c == "  ":
                value = {"code": "l+",
                         "status": "insertLine",
                         "from": "",
                         "to": to_dict[key_v][column_v],
                         "value": to_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)
            elif key_c == "  " and column_c == "+ ":
                value = {"code": "r+",
                         "status": "insertRow",
                         "from": "",
                         "to": to_dict[key_v][column_v],
                         "value": to_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)
            elif key_c == "+ " and column_c == "+ ":
                value = {"code": "++",
                         "status": "insert",
                         "from": "",
                         "to": to_dict[key_v][column_v],
                         "value": to_dict[key_v][column_v]
                         }
                diff_results[key_v].setdefault(column_v, value)

    results = {"diff_results": diff_results,
               "diff_keys": diff_keys,
               "diff_columns": diff_columns,
               "delete_rows": helpers.list_difference(from_keys, to_keys),
               "delete_columns": helpers.list_difference(from_columns, to_columns),
               "insert_rows": helpers.list_difference(to_keys, from_keys),
               "insert_columns": helpers.list_difference(to_columns, from_columns)
               }
    return results


def show_diff_excel_data_as_text(diff_excel_data_results: dict) -> str:
    results = []

    if diff_excel_data_results['delete_rows']:
        results.append("{0}行が削除されました．".format(len(diff_excel_data_results['delete_rows'])))
        for i in diff_excel_data_results['delete_rows']:
            results.append("  行番号: {0}".format(i))
    if diff_excel_data_results['delete_columns']:
        results.append("{0}列が削除されました．".format(len(diff_excel_data_results['delete_columns'])))
        for i in diff_excel_data_results['delete_columns']:
            results.append("  列番号: {0}".format(i))

    if diff_excel_data_results['insert_rows']:
        results.append("{0}行が追加されました．".format(len(diff_excel_data_results['insert_rows'])))
        for i in diff_excel_data_results['insert_rows']:
            results.append("  行番号: {0}".format(i))

    if diff_excel_data_results['insert_columns']:
        results.append("{0}列が追加されました．".format(len(diff_excel_data_results['insert_columns'])))
        for i in diff_excel_data_results['insert_columns']:
            results.append("  列番号: {0}".format(i))
    if len(results) == 0:
        results.append("行と列の削除や追加はありませんでした．")

    results.append("======================================")

    current_row = ""

    diff_results = diff_excel_data_results["diff_results"]
    diff_keys = diff_excel_data_results["diff_keys"]
    diff_columns = diff_excel_data_results["diff_columns"]
    # print(diff_results)
    for r in diff_keys:
        u = diff_results.get(r)
        if u is None:
            raise Exception("行データが見当たらない")
        for c in diff_columns:
            v = u.get(c)
            if v is None:
                continue
            if "-" not in v["code"]:
                if "replace" == v["status"]:
                    if current_row != r:
                        results.append("{}行: ".format(r))
                        current_row = r
                    results.append("  {}列: {} -> {}".format(c, v["from"], v["to"]))
                # elif "+" in v["code"]:
                #     if current_row != r:
                #         results.append("{}行: ".format(r))
                #         current_row = r
                #     results.append("  {}列: [NEW] {}".format(c, v["to"]))

    results.append("======================================")

    return "\n".join(results)


def show_diff_excel_data_as_xlsx(diff_excel_data_results: dict) -> None:
    unchange = NamedStyle(name="unchange")
    unchange.font = Font(color='FF000000')
    unchange.fill = PatternFill()

    delete = NamedStyle(name="delete")
    delete.font = Font(color='FF000000')
    delete.fill = PatternFill(fill_type='solid', fgColor='fffff4e0')

    insert = NamedStyle(name="insert")
    insert.font = Font(color='FF000000')
    insert.fill = PatternFill(fill_type='solid', fgColor='ffeeffeb')

    insert_delete = NamedStyle(name="insert_delete")
    insert_delete.font = Font(color='FF000000')
    insert_delete.fill = PatternFill(fill_type='solid', fgColor='fff9fae1')

    replace = NamedStyle(name="replace")
    replace.font = Font(color='FF000000')
    replace.fill = PatternFill(fill_type='solid', fgColor='ffdceefe')

    label = NamedStyle(name="label")
    label.font = Font(color='FF000000', b=True)
    label.fill = PatternFill(fill_type='solid', fgColor="ffdfdfdf")

    label_delete = NamedStyle(name="label_delete")
    label_delete.font = Font(color='FF000000', b=True)
    label_delete.fill = PatternFill(fill_type='solid', fgColor="ffd9a5a3")

    label_insert = NamedStyle(name="label_insert")
    label_insert.font = Font(color='FF000000', b=True)
    label_insert.fill = PatternFill(fill_type='solid', fgColor="ffadccae")

    wb = openpyxl.Workbook()

    wb.add_named_style(unchange)
    wb.add_named_style(delete)
    wb.add_named_style(insert)
    wb.add_named_style(insert_delete)
    wb.add_named_style(replace)
    wb.add_named_style(label)
    wb.add_named_style(label_delete)
    wb.add_named_style(label_insert)
    ws = wb.active
    ws.title = "results"

    diff_results = diff_excel_data_results["diff_results"]
    diff_keys = diff_excel_data_results["diff_keys"]
    diff_columns = diff_excel_data_results["diff_columns"]

    for i, r in enumerate(diff_keys):
        u = diff_results.get(r)
        if u is None:
            raise Exception("行データが見当たらない")

        for j, c in enumerate(diff_columns):
            v = u.get(c, {"value": "", "code": None})
            if v["code"] is None:
                style = insert_delete
            elif "-" in v["code"]:
                style = delete
            elif ">" in v["code"]:
                style = replace
            elif "+" in v["code"]:
                style = insert
            else:
                style = unchange

            _ = ws.cell(row=i + 2, column=j + 2, value=v["value"])
            _.style = style

    _ = ws.cell(row=1, column=1, value="")
    _.style = label

    for i, r in enumerate(diff_keys):
        _ = ws.cell(row=i + 2, column=1, value="{}".format(r))
        if r in diff_excel_data_results['delete_rows']:
            _.style = label_delete
        elif r in diff_excel_data_results['insert_rows']:
            _.style = label_insert
        else:
            _.style = label

    for j, c in enumerate(diff_columns):
        _ = ws.cell(row=1, column=j + 2, value="{}".format(c))
        if c in diff_excel_data_results['delete_columns']:
            _.style = label_delete
        elif c in diff_excel_data_results['insert_columns']:
            _.style = label_insert
        else:
            _.style = label

    date = dt.now().astimezone(pytz.timezone("Asia/Tokyo")).strftime('%Y%m%dT%H%M%S')
    output_file_path = Path('diff_results_{}.xlsx'.format(date))
    wb.save(output_file_path.resolve())
    print("{0} に結果を保存しました．".format(output_file_path.resolve()))
